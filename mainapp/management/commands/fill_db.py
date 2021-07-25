from django.core.management.base import BaseCommand
# from django.contrib.auth.models import User
# from django.contrib.auth.models.User(AbstractUser) import User

import json
import os

from mainapp.models import ProductCategory, Product
from authapp.models import ShopUser

from openpyxl import Workbook   # для работы с Excel
import pandas as pd             # для работы с Excel

# from django.core.exceptions import DoesNotExist     # для обработки исключения


JSON_PATH = 'mainapp/import'
EXCEL_PATH = 'mainapp/import'

class Command(BaseCommand):
    def handle(self, *args, **options):
        # fill_new_bd_from_EXCEL_file('catalog')   # Зальем новую БД с перечнем товаров и категорий (запускаем при изменении Excel-файла)

        # categories = load_from_json('categories')
        # ProductCategory.objects.all().delete()
        # for category in categories:
        #     new_category = ProductCategory(**category)
        #     new_category.save()
        #
        # products = load_from_json('products')
        # Product.objects.all().delete()
        # for product in products:
        #     category_name = product["category"]
        #     # Получаем категорию по имени
        #     _category = ProductCategory.objects.get(name=category_name)
        #     # Заменяем название категории объектом
        #     product['category'] = _category
        #     new_product = Product(**product)
        #     new_product.save()

        # Создаем суперпользователя при помощи менеджера модели
        ShopUser.objects.create_superuser('kuznetsov', 'ksn1974@mail.ru', '1', age=33)

def load_from_json(file_name):
    with open(os.path.join(JSON_PATH, file_name + '.json'), 'r') as infile:
        return json.load(infile)

################################################
# Вспомогательные функции для работы с файлами
################################################

# Функция загрузки БД из EXCEL-файла (его проще набивать) текущая БД удаляется
def fill_new_bd_from_EXCEL_file (file_name="catalog"):
    convert_EXCEL_file_to_JSON_file(file_name)      # Конвертируем перечень товаров из excel (легче набивать) в JSON
    catalog = read_json(file_name)                  # Считаем перечень из JSON

    ProductCategory.objects.all().delete()      # Удалим все из справочников ProductCategory и Product (каскадно)
    # for el in ProductCategory.objects.all():    # Удалим все из справочников ProductCategory и Product (каскадно)
    #     el.delete()

    for el in catalog:
        ProductCategory.objects.get_or_create(name=el['category'])      # Создадим новую категорию, если такой не было
        category = ProductCategory.objects.filter(name=el['category'])  # Найдем эту категорию для получения ID
        product = Product(name=el['name'], category_id=category.values()[0]['id'], short_desc=el['short_desc'], image='products_images/'+el['image'], description=el['description'], price=el['price'], quantity=el['quantity'])
        product.save()

# Конвертируем перечень товаров из excel в JSON
def convert_EXCEL_file_to_JSON_file (file_name):
    # data = pd.read_excel(path, usecols="A,E:F", encoding='utf8')
    data = pd.read_excel(os.path.join(EXCEL_PATH, file_name + '.xlsx'))
    # i = path.find('.', len(path)-6)     # избавимся от расширения
    # path_JSON = path[:i]+'.json'         # установим расширение файла для сохранения
    path_JSON = os.path.join(JSON_PATH, file_name + '.json')
    data.to_json(path_or_buf=path_JSON, orient='records')
    return path_JSON

# считаем JSON из файла
def read_json(file_name):
    with open(os.path.join(JSON_PATH, file_name + '.json'), 'r') as infile:
        return json.load(infile)
    # return json.loads(read_file(file_name_with_path))

#**********************************
# остальное пока не нужно....
#**********************************

def read_file(path):
    file = open(path, "r")
    data = file.read()
    file.close()
    return data

def write_json(path, data):
    return write_file(path, json.dumps(data))

def write_excel(file_name="sample", data='Ерунда'):
    import datetime         # Python types will automatically be converted
    wb = Workbook()
    ws = wb.active          # grab the active worksheet
    ws['A1'] = data         # Data can be assigned directly to cells
    ws.append([1, 2, 3])    # Rows can also be appended
    ws['A2'] = datetime.datetime.now()
    return wb.save(os.path.join(EXCEL_PATH, file_name + '.xlsx'))   # Save the file

def write_file(path, data):
    file = open(path, "w")
    file.write(str(data))
    file.close()
    return data

# Функция записи в первый раз в JSON-файл перечня меню
def write_links_menu_to_JSON_file ():
    json_data = [
        {'href': 'index', 'name': 'все'},
        {'href': 'products_home', 'name': 'дом'},
        {'href': 'products_office', 'name': 'офис'},
        {'href': 'products_modern', 'name': 'модерн'},
        {'href': 'products_classic', 'name': 'классика'},
    ]
    write_json('./geekshop/links_menu.JSON', json_data);

def read_JSON_from_Excel_file (file_name):
    # data = pd.read_excel(path, usecols="A,E:F", encoding='utf8')
    data = pd.read_excel(os.path.join(EXCEL_PATH, file_name + '.xlsx'))
    # return json.loads(data.to_json())
    # return data.to_json()
    return data.to_dict()




