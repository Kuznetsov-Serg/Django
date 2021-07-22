from django.shortcuts import render
import json
from .models import ProductCategory, Product
from openpyxl import Workbook   # дЗля работы с Excel
import pandas as pd             # для работы с Excel
# from django.core.exceptions import DoesNotExist     # для обработки исключения
from django.urls import reverse


# Create your views here.

# def index(request):
#     return render(request, 'mainapp/products.html')
#
# def products(request):
#     return render(request, 'mainapp/products.html')
def get_absolute_url(self):
    return reverse('products', kwargs={'category_id': self.category_id})

def index(request, category_id=0):
    # category_id - вх. параметр для страницы products (фильтр для показа related products) если=0 - пункт "ВСЕ"

    # write_links_menu_to_JSON_file()
    # fill_new_bd_from_EXCEL_file()               # Зальем новую БД с перечнем товаров и категорий (запускаем при изменении Excel-файла)

    title = 'каталог'
    # Вариант статической загрузки категорий
    # links_menu = [
    #     {'href': 'index', 'name': 'все!!!'},
    #     {'href': 'products_home', 'name': 'дом'},
    #     {'href': 'products_office', 'name': 'офис'},
    #     {'href': 'products_modern', 'name': 'модерн'},
    #     {'href': 'products_classic', 'name': 'классика'},
    # ]

    # Вариант загрузки категорий из JSON-файла
    # links_menu = read_json('./links_menu.JSON')     # считаем перечень меню из JSON-файла (раньше)

    # Вариант загрузки категорий из БД (таблица ProductsCategory)
    # links_menu = ProductCategory.objects.all()      # Считаем все категории из справочника (не получим "все")
    links_menu = [{'href': '/products0/', 'id': 0, 'name': 'все'}]  # дополним пунктом меню для всех категорий
    for el in ProductCategory.objects.all():                        # Считаем все категории из справочника
        links_menu.append({'href': '/products'+str(el.id)+'/', 'id': el.id, 'name': el.name})

    # products = Product.objects.all()[:3]                # Считаем продукты из справочника
    # products = Product.objects.all()    # Считаем продукты из справочника (количество уже ограничим в форме)
    products = []
    products_amount = {}
    for el in Product.objects.all():
        if products_amount.get(el.category_id) == None:     # не было такой категории
            products_amount[el.category_id] = 0
        if products_amount[el.category_id] < 3:             # берем не более трех товаров каждой категории
            products_amount[el.category_id] += 1
            products.append(el)

    # print(f'category_id= {category_id}')
    context = {
        'title': title,
        'links_menu': links_menu,
        'related_products': products,
        'category_id': category_id,
    }
    return render(request, 'mainapp/products.html', context)

################################################
# Вспомогательные функции для работы с файлами
################################################

def read_file(path):
    file = open(path, "r")
    data = file.read()
    file.close()
    return data

def read_json(path):
    return json.loads(read_file(path))

def write_json(path, data):
    return write_file(path, json.dumps(data))

def write_excel(path="sample.xlsx", data='Ерунда'):
    import datetime         # Python types will automatically be converted
    wb = Workbook()
    ws = wb.active          # grab the active worksheet
    ws['A1'] = data         # Data can be assigned directly to cells
    ws.append([1, 2, 3])    # Rows can also be appended
    ws['A2'] = datetime.datetime.now()
    return wb.save(path)   # Save the file

def write_file(path, data):
    file = open(path, "w")
    file.write(str(data))
    file.close()
    return data

# Функция записи в первый раз в JSON-файл перечня меню
def write_links_menu_to_JSON_file ():
    json_data = [
        {'href': 'index', 'name': 'все'},
        {'href': 'products_home', 'name': 'дом'},
        {'href': 'products_office', 'name': 'офис'},
        {'href': 'products_modern', 'name': 'модерн'},
        {'href': 'products_classic', 'name': 'классика'},
    ]
    write_json('./geekshop/links_menu.JSON', json_data);

def read_JSON_from_Excel_file (path="products.xlsx"):
    # data = pd.read_excel(path, usecols="A,E:F", encoding='utf8')
    data = pd.read_excel(path)
    # return json.loads(data.to_json())
    # return data.to_json()
    return data.to_dict()

def convert_EXCEL_file_to_JSON_file (path="products.xlsx"):
    # data = pd.read_excel(path, usecols="A,E:F", encoding='utf8')
    data = pd.read_excel(path)
    i = path.find('.', len(path)-6)     # избавимся от расширения
    path_JSON = path[:i]+'.json'         # установим расширение файла для сохранения
    data.to_json(path_or_buf=path_JSON, orient='records')
    return path_JSON

def fill_new_bd_from_EXCEL_file (path="products.xlsx"):
    path_JSON = convert_EXCEL_file_to_JSON_file(path)   # Конвертируем перечень товаров из excel (легче набивать) в JSON
    catalog = read_json(path_JSON)                      # Считаем каталог

    for el in ProductCategory.objects.all():    # Удалим все из справочников ProductCategory и Product (каскадно)
        el.delete()
    # ttt = ProductCategory()
    # products = Product.objects.all()
    # product_categorys = ProductCategory.objects.all()
    for el in catalog:
        # print(el)
        # print(el['category'])
        ttt = ProductCategory.objects.get_or_create(name=el['category'])        # Создадим новую категорию, если такой не было
        rrr = ProductCategory.objects.filter(name=el['category'])               # Найдем эту категорию для получения ID
        mmm = Product(name=el['name'], category_id=rrr.values()[0]['id'], short_desc=el['short_desc'], image='products_images/'+el['image'], description=el['description'], price=el['price'], quantity=el['quantity'])
        mmm.save()

